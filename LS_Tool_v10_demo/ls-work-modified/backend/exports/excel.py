from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date
from collections import defaultdict

# ── Palette ───────────────────────────────────────────────────────────────────
PRIMARY    = "1E50A0"
HEADER2    = "4472C4"
LIGHT_BLUE = "DCE6F5"
WHITE      = "FFFFFF"
GREEN_F    = "E2EFDA"; GREEN_T  = "375623"
AMBER_F    = "FFF2CC"; AMBER_T  = "7F6000"
RED_F      = "FCE4D6"; RED_T    = "9C0006"
PURPLE_F   = "EAE7F7"; PURPLE_T = "3B1F8C"
GRAY_F     = "F2F2F2"

ZAR_FMT  = 'R#,##0.00'
PCT_FMT  = '0.0%'
NUM_FMT  = '#,##0.0'

# ── Style helpers ─────────────────────────────────────────────────────────────
def hdr(ws, row, cols, color=PRIMARY):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def section(ws, text, color=PRIMARY):
    ws.append([text])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color=color)
    ws.append([])

def paint(ws, row, col, fill, font_color=None, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.fill = PatternFill("solid", fgColor=fill)
    kw = {}
    if font_color: kw["color"] = font_color
    if bold:       kw["bold"] = bold
    if kw:         cell.font = Font(**kw)

def zar(ws, row, cols):
    for c in cols:
        ws.cell(row=row, column=c).number_format = ZAR_FMT

def auto_width(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value: mx = max(mx, len(str(cell.value)))
            except: pass
        ws.column_dimensions[letter].width = min(mx + 4, 52)

def zebra(ws, r0, r1, n_cols):
    for r in range(r0, r1 + 1):
        if r % 2 == 0:
            for c in range(1, n_cols + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

def bold_row(ws, row, n_cols, fill=LIGHT_BLUE):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=fill)

# ── Maths helpers ─────────────────────────────────────────────────────────────
def linreg(ys):
    """Return (slope, intercept) for a list of y-values."""
    n = len(ys)
    if n < 2: return 0, ys[0] if ys else 0
    xm = (n - 1) / 2.0
    ym = sum(ys) / n
    num = sum((i - xm) * (v - ym) for i, v in enumerate(ys))
    den = sum((i - xm) ** 2 for i in range(n))
    s = num / den if den else 0
    return s, ym - s * xm

def rolling_avg(series, window=3):
    return sum(series[-window:]) / min(window, len(series)) if series else 0

def forecast(series, n=3, weight_linear=0.6):
    """Blend linear regression and rolling average to project n periods ahead."""
    if not series: return [0] * n
    slope, intercept = linreg(series)
    ra = rolling_avg(series)
    base = len(series)
    out = []
    for i in range(n):
        lin = intercept + slope * (base + i)
        out.append(max(lin * weight_linear + ra * (1 - weight_linear), 0))
    return out

def add_months(year, month, n=1):
    total = (year - 1) * 12 + (month - 1) + n
    return total // 12 + 1, total % 12 + 1

MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def mlabel(year, month):
    return f"{MONTH_NAMES[month-1]} {year}"

def money(v):
    return f"R{v:,.2f}" if v is not None else "R0.00"

# ── Main export ───────────────────────────────────────────────────────────────
def generate_full_report(db) -> BytesIO:
    from models import (Order, Payment, Delivery, TLBBooking,
                        Product, PriceTier, JournalEntry, Account)
    from gl.service import get_trial_balance, get_profit_loss, get_ar_aging
    from sqlalchemy import func

    wb   = Workbook()
    today = date.today()

    # ── Pre-load ──────────────────────────────────────────────────────────────
    orders      = db.query(Order).order_by(Order.order_date.asc()).all()
    active      = [o for o in orders if o.status != "cancelled"]
    payments    = db.query(Payment).order_by(Payment.payment_date.desc()).all()
    deliveries  = db.query(Delivery).all()
    tlbs        = db.query(TLBBooking).all()
    total_rev   = sum(o.total_zar for o in active)
    total_paid  = sum(p.amount_zar for p in payments)

    # Monthly buckets
    monthly = {}
    for o in active:
        k = o.order_date.strftime("%Y-%m")
        if k not in monthly:
            monthly[k] = {"orders": 0, "revenue": 0.0, "items": 0,
                          "yr": o.order_date.year, "mo": o.order_date.month}
        monthly[k]["orders"]  += 1
        monthly[k]["revenue"] += o.total_zar
        monthly[k]["items"]   += sum(i.quantity for i in o.items)
    months_sorted = sorted(monthly.keys())

    # Product demand buckets
    prod = defaultdict(lambda: {
        "qty": 0, "revenue": 0.0, "orders": set(),
        "cat": "", "last": None,
        "by_month": defaultdict(int)
    })
    for o in active:
        mk = o.order_date.strftime("%Y-%m")
        for item in o.items:
            p = prod[item.product.name]
            p["qty"]      += item.quantity
            p["revenue"]  += item.line_total_zar
            p["orders"].add(o.id)
            p["cat"]       = item.product.category
            p["by_month"][mk] += item.quantity
            if not p["last"] or o.order_date > p["last"]:
                p["last"] = o.order_date

    # Customer RFM buckets
    cust = defaultdict(lambda: {
        "orders": 0, "revenue": 0.0,
        "first": None, "last": None
    })
    for o in active:
        c = cust[o.customer.name]
        c["orders"]  += 1
        c["revenue"] += o.total_zar
        if not c["first"] or o.order_date < c["first"]: c["first"] = o.order_date
        if not c["last"]  or o.order_date > c["last"]:  c["last"]  = o.order_date

    n_months = max(len(months_sorted), 1)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active; ws.title = "Summary"
    ws.append(["LOGISTICS SYSTEM — FULL REPORT"])
    ws["A1"].font = Font(bold=True, size=16, color=PRIMARY)
    ws.append([f"Generated: {today.strftime('%d %B %Y')}"])
    ws.append([])
    ws.append(["Metric", "Value"])
    hdr(ws, 4, 2)
    rows = [
        ("Total Revenue (active orders)",   total_rev),
        ("Total Payments Received",          total_paid),
        ("Outstanding Balance",              total_rev - total_paid),
        ("Total Orders",                     len(orders)),
        ("Active Customers",                 len(cust)),
        ("Products in Catalogue",            db.query(Product).filter_by(is_active=True).count()),
        ("Deliveries Completed",             db.query(Delivery).filter_by(status="completed").count()),
        ("TLB Bookings",                     len(tlbs)),
        ("Lay-buys Active",                  0),  # placeholder — model may not exist yet
    ]
    # Try to get laybuy count if model exists
    try:
        from models import Laybuy
        rows[-1] = ("Lay-buys Active", db.query(Laybuy).filter_by(status="active").count())
    except Exception:
        pass

    for i, (label, val) in enumerate(rows, 5):
        ws.append([label, val])
        if isinstance(val, float):
            ws.cell(i, 2).number_format = ZAR_FMT
    zebra(ws, 5, 4 + len(rows), 2)
    auto_width(ws)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Orders
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Orders")
    h2 = ["Order No.", "Customer", "Date", "Status",
          "Items Subtotal", "Transport", "Disc %", "Disc Amt",
          "Disc Reason", "Total", "Paid", "Balance", "Notes"]
    ws2.append(h2); hdr(ws2, 1, len(h2))
    for i, o in enumerate(reversed(orders), 2):
        paid = sum(p.amount_zar for p in o.payments)
        sub  = sum(item.line_total_zar for item in o.items)
        ws2.append([o.order_number, o.customer.name, str(o.order_date), o.status,
                    sub, o.transport_price or 0, f"{o.discount_pct or 0:.1f}%",
                    o.discount_amount or 0, o.discount_reason or "",
                    o.total_zar, paid, o.total_zar - paid, o.notes or ""])
        zar(ws2, i, [5, 6, 8, 10, 11, 12])
    zebra(ws2, 2, len(orders) + 1, len(h2)); auto_width(ws2)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Order Items
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Order Items")
    h3 = ["Order No.", "Customer", "Date", "Product", "Category",
          "Tier", "Tier Price", "Custom?", "Unit Price", "Qty", "Line Total", "Reason"]
    ws3.append(h3); hdr(ws3, 1, len(h3))
    ri = 2
    for o in reversed(orders):
        for item in o.items:
            ws3.append([o.order_number, o.customer.name, str(o.order_date),
                        item.product.name, item.product.category,
                        item.price_tier.tier_label, item.price_tier.price_zar,
                        "Yes" if item.custom_unit_price is not None else "No",
                        item.unit_price_zar, item.quantity, item.line_total_zar,
                        item.custom_price_reason or ""])
            zar(ws3, ri, [7, 9, 11]); ri += 1
    zebra(ws3, 2, ri, len(h3)); auto_width(ws3)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Deliveries
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Deliveries")
    h4 = ["Order No.", "Customer", "Type", "Hardware Store", "Scheduled", "Actual",
          "Driver", "Vehicle", "Status", "Live Location Link"]
    ws4.append(h4); hdr(ws4, 1, len(h4))
    for i, d in enumerate(deliveries, 2):
        ws4.append([d.order.order_number, d.order.customer.name,
                    d.delivery_type or "customer",
                    getattr(d, "hardware_store_name", None) or "",
                    str(d.scheduled_date), str(d.actual_date) if d.actual_date else "",
                    d.driver_name or "", d.vehicle_reg or "", d.status,
                    d.live_location_link or ""])
    zebra(ws4, 2, len(deliveries) + 1, len(h4)); auto_width(ws4)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — TLB Jobs
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("TLB Jobs")
    h5 = ["ID", "Customer", "Date", "Hours", "Rate Type", "Custom Rate?",
          "Effective Rate", "Rate Reason", "Diesel", "Transport",
          "Transport Notes", "Total", "Status"]
    ws5.append(h5); hdr(ws5, 1, len(h5))
    for i, t in enumerate(tlbs, 2):
        if t.rate_type == "hourly":
            er = t.custom_hourly_rate if t.custom_hourly_rate is not None else 1000.0
            ic = "Yes" if t.custom_hourly_rate is not None else "No"
        else:
            er = t.custom_daily_rate if t.custom_daily_rate is not None else 7500.0
            ic = "Yes" if t.custom_daily_rate is not None else "No"
        ws5.append([f"TLB-{t.id}", t.customer.name, str(t.booking_date),
                    t.hours_billed, t.rate_type, ic, er, t.custom_rate_reason or "",
                    "Yes" if t.diesel_included else "No",
                    t.transport_price or 0, t.transport_notes or "",
                    t.total_zar, t.status])
        zar(ws5, i, [7, 10, 12])
    zebra(ws5, 2, len(tlbs) + 1, len(h5)); auto_width(ws5)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 6 — Payments
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Payments")
    h6 = ["ID", "Order / TLB", "Amount", "Date", "Method", "Reference"]
    ws6.append(h6); hdr(ws6, 1, len(h6))
    for i, p in enumerate(payments, 2):
        ws6.append([p.id,
                    p.order.order_number if p.order else f"TLB-{p.tlb_booking_id}",
                    p.amount_zar, str(p.payment_date), p.method, p.reference or ""])
        zar(ws6, i, [3])
    zebra(ws6, 2, len(payments) + 1, len(h6)); auto_width(ws6)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 7 — Price List
    # ══════════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Price List")
    h7 = ["Product", "Category", "Tier", "Quantity", "Price"]
    ws7.append(h7); hdr(ws7, 1, len(h7))
    tiers = (db.query(PriceTier)
             .filter(PriceTier.effective_to == None)
             .join(Product).filter(Product.is_active == True).all())
    for i, t in enumerate(tiers, 2):
        ws7.append([t.product.name, t.product.category, t.tier_label, t.quantity, t.price_zar])
        zar(ws7, i, [5])
    zebra(ws7, 2, len(tiers) + 1, len(h7)); auto_width(ws7)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 8 — Trial Balance
    # ══════════════════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("Trial Balance")
    ws8.append([f"TRIAL BALANCE — {today.strftime('%d %B %Y')}"])
    ws8["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws8.append([])
    h8 = ["Code", "Account Name", "Type", "Debit", "Credit", "Net"]
    ws8.append(h8); hdr(ws8, 3, len(h8))
    tb = get_trial_balance(db)
    td = tc = 0
    for i, r in enumerate(tb, 4):
        ws8.append([r["code"], r["name"], r["account_type"],
                    r["total_debit"] or 0, r["total_credit"] or 0, r["net"]])
        zar(ws8, i, [4, 5, 6])
        td += r["total_debit"]; tc += r["total_credit"]
    ws8.append([]); ws8.append(["", "TOTALS", "", td, tc, ""])
    last = ws8.max_row
    bold_row(ws8, last, len(h8))
    zar(ws8, last, [4, 5])
    zebra(ws8, 4, last - 2, len(h8)); auto_width(ws8)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 9 — Profit & Loss
    # ══════════════════════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("Profit & Loss")
    ws9.append([f"PROFIT & LOSS — {today.strftime('%d %B %Y')}"])
    ws9["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws9.append([])
    pl = get_profit_loss(db)
    section(ws9, "REVENUE")
    ws9.append(["Code", "Account", "Amount"]); hdr(ws9, ws9.max_row, 3)
    for r in pl["revenue"]:
        ws9.append([r["code"], r["name"], r["amount"]])
        zar(ws9, ws9.max_row, [3])
    ws9.append(["", "Total Revenue", pl["total_revenue"]])
    ws9.cell(ws9.max_row, 2).font = Font(bold=True)
    zar(ws9, ws9.max_row, [3])
    ws9.append([])
    section(ws9, "EXPENSES")
    ws9.append(["Code", "Account", "Amount"]); hdr(ws9, ws9.max_row, 3)
    for e in pl["expenses"]:
        ws9.append([e["code"], e["name"], e["amount"]])
        zar(ws9, ws9.max_row, [3])
    ws9.append(["", "Total Expenses", pl["total_expenses"]])
    ws9.cell(ws9.max_row, 2).font = Font(bold=True)
    zar(ws9, ws9.max_row, [3])
    ws9.append([])
    ws9.append(["", "NET PROFIT / (LOSS)", pl["net_profit"]])
    net_cell = ws9.cell(ws9.max_row, 3)
    net_cell.number_format = ZAR_FMT
    net_cell.font = Font(bold=True, size=12,
                         color=GREEN_T if pl["net_profit"] >= 0 else RED_T)
    ws9.cell(ws9.max_row, 2).font = Font(bold=True, size=12)
    auto_width(ws9)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 10 — AR Aging
    # ══════════════════════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("AR Aging")
    ws10.append([f"ACCOUNTS RECEIVABLE AGING — {today.strftime('%d %B %Y')}"])
    ws10["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws10.append([])
    h10 = ["Customer","Order No.","Date","Total","Paid","Balance",
           "Current (0-30d)","31-60d","61-90d","90d+"]
    ws10.append(h10); hdr(ws10, 3, len(h10))
    aging = get_ar_aging(db)
    for i, r in enumerate(aging, 4):
        ws10.append([r["customer"], r["order_number"], r["order_date"],
                     r["total"], r["paid"], r["balance"],
                     r["current"], r["days_31_60"], r["days_61_90"], r["days_90_plus"]])
        zar(ws10, i, [4, 5, 6, 7, 8, 9, 10])
        if r["days_90_plus"] > 0: paint(ws10, i, 10, RED_F,   RED_T)
        if r["days_61_90"] > 0:   paint(ws10, i,  9, AMBER_F, AMBER_T)
    zebra(ws10, 4, 3 + len(aging), len(h10)); auto_width(ws10)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 11 — General Ledger
    # ══════════════════════════════════════════════════════════════════════════
    ws11 = wb.create_sheet("General Ledger")
    h11 = ["Date", "Ref", "Account Code", "Account", "Description", "Debit", "Credit"]
    ws11.append(h11); hdr(ws11, 1, len(h11))
    entries = (db.query(JournalEntry)
               .order_by(JournalEntry.entry_date, JournalEntry.transaction_ref).all())
    for i, e in enumerate(entries, 2):
        ws11.append([str(e.entry_date), e.transaction_ref,
                     e.account.code, e.account.name, e.description,
                     e.debit_zar or 0, e.credit_zar or 0])
        zar(ws11, i, [6, 7])
    zebra(ws11, 2, len(entries) + 1, len(h11)); auto_width(ws11)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 12 — Monthly Revenue Trend
    # ══════════════════════════════════════════════════════════════════════════
    ws12 = wb.create_sheet("Monthly Trend")
    ws12.append(["MONTHLY REVENUE & VOLUME TREND"])
    ws12["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws12.append([f"Covers {len(months_sorted)} months · {len(active)} active orders · "
                 f"Use this sheet to spot seasonality and growth/decline patterns"])
    ws12.cell(2, 1).font = Font(italic=True, color="666666")
    ws12.append([])

    h12 = ["Month", "# Orders", "Revenue", "Avg Order Value", "Units Sold",
           "MoM Revenue Δ", "MoM Δ %", "3-Mo Rolling Avg Revenue", "Cumulative Revenue"]
    ws12.append(h12); hdr(ws12, 4, len(h12))

    rev_series = [monthly[k]["revenue"] for k in months_sorted]
    cumul = 0.0
    prev_rev = None
    win = []
    for ri, k in enumerate(months_sorted, 5):
        m = monthly[k]
        rev  = m["revenue"]
        n_ord = m["orders"]
        avg  = rev / n_ord if n_ord else 0
        cumul += rev
        win.append(rev)
        if len(win) > 3: win.pop(0)
        ra = sum(win) / len(win)

        mom_d = (rev - prev_rev) if prev_rev is not None else None
        mom_p = (mom_d / prev_rev) if (prev_rev and mom_d is not None) else None

        ws12.append([mlabel(m["yr"], m["mo"]), n_ord, rev, avg, m["items"],
                     mom_d if mom_d is not None else "",
                     mom_p if mom_p is not None else "",
                     ra, cumul])
        zar(ws12, ri, [3, 4, 6, 8, 9])
        if mom_p is not None:
            ws12.cell(ri, 7).number_format = PCT_FMT
            if   mom_p >  0.1: paint(ws12, ri, 7, GREEN_F, GREEN_T)
            elif mom_p < -0.1: paint(ws12, ri, 7, RED_F,   RED_T)
        prev_rev = rev

    # Totals
    ws12.append([])
    ws12.append(["TOTAL / AVERAGE",
                 sum(m["orders"] for m in monthly.values()),
                 total_rev,
                 total_rev / len(active) if active else 0,
                 sum(m["items"] for m in monthly.values()),
                 "", "", "", total_rev])
    bold_row(ws12, ws12.max_row, len(h12))
    zar(ws12, ws12.max_row, [3, 4, 9])
    zebra(ws12, 5, 4 + len(months_sorted), len(h12))
    auto_width(ws12)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 12b — Hardware Channel (hardware-store revenue vs direct)
    # ══════════════════════════════════════════════════════════════════════════
    def _is_hw(o):
        if o.customer and getattr(o.customer, "customer_type", None) == "hardware_store":
            return True
        d_ = o.delivery
        return bool(d_ and d_.delivery_type == "hardware")

    hw_orders = [o for o in orders if o.status in ("confirmed", "delivered")]
    hw_monthly = {}
    for o in hw_orders:
        k = o.order_date.strftime("%Y-%m")
        if k not in hw_monthly:
            hw_monthly[k] = {"hw_n": 0, "hw_rev": 0.0, "dr_n": 0, "dr_rev": 0.0,
                             "yr": o.order_date.year, "mo": o.order_date.month}
        if _is_hw(o):
            hw_monthly[k]["hw_n"] += 1; hw_monthly[k]["hw_rev"] += o.total_zar
        else:
            hw_monthly[k]["dr_n"] += 1; hw_monthly[k]["dr_rev"] += o.total_zar

    wshc = wb.create_sheet("Hardware Channel")
    wshc.append(["HARDWARE CHANNEL — REVENUE BY MONTH"])
    wshc["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    wshc.append(["Hardware = orders placed by a hardware-store account or delivered to a hardware store. "
                 "Confirmed/delivered orders only (quotes & cancellations excluded)."])
    wshc.cell(2, 1).font = Font(italic=True, color="666666")
    wshc.append([])
    hhc = ["Month", "Hardware Orders", "Hardware Revenue", "Direct Orders",
           "Direct Revenue", "Total Revenue", "Hardware Share %"]
    wshc.append(hhc); hdr(wshc, 4, len(hhc))
    hw_keys = sorted(hw_monthly.keys())
    for ri, k in enumerate(hw_keys, 5):
        m = hw_monthly[k]
        total = m["hw_rev"] + m["dr_rev"]
        share = (m["hw_rev"] / total) if total else 0
        wshc.append([mlabel(m["yr"], m["mo"]), m["hw_n"], m["hw_rev"],
                     m["dr_n"], m["dr_rev"], total, share])
        zar(wshc, ri, [3, 5, 6])
        wshc.cell(ri, 7).number_format = PCT_FMT
    wshc.append([])
    tot_hw = sum(m["hw_rev"] for m in hw_monthly.values())
    tot_dr = sum(m["dr_rev"] for m in hw_monthly.values())
    tot_all = tot_hw + tot_dr
    wshc.append(["TOTAL", sum(m["hw_n"] for m in hw_monthly.values()), tot_hw,
                 sum(m["dr_n"] for m in hw_monthly.values()), tot_dr, tot_all,
                 (tot_hw / tot_all) if tot_all else 0])
    bold_row(wshc, wshc.max_row, len(hhc))
    zar(wshc, wshc.max_row, [3, 5, 6])
    wshc.cell(wshc.max_row, 7).number_format = PCT_FMT
    if hw_keys:
        zebra(wshc, 5, 4 + len(hw_keys), len(hhc))
    auto_width(wshc)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 13 — Product Demand
    # ══════════════════════════════════════════════════════════════════════════
    ws13 = wb.create_sheet("Product Demand")
    ws13.append(["PRODUCT DEMAND ANALYSIS"])
    ws13["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws13.append(["Units sold, revenue velocity and trend per product. "
                 "Sort any column. Pivot below shows monthly units per product."])
    ws13.cell(2, 1).font = Font(italic=True, color="666666")
    ws13.append([])

    h13 = ["Product", "Category", "Total Units", "Total Revenue",
           "Avg Unit Price", "# Orders", "Units/Month (avg)",
           "Revenue Share", "Last Ordered", "Trend"]
    ws13.append(h13); hdr(ws13, 4, len(h13))

    prod_sorted = sorted(prod.items(), key=lambda x: x[1]["revenue"], reverse=True)

    for ri, (pname, pd) in enumerate(prod_sorted, 5):
        n_ord  = len(pd["orders"])
        avg_px = pd["revenue"] / pd["qty"] if pd["qty"] else 0
        avg_mo = pd["qty"] / n_months
        rs     = pd["revenue"] / total_rev if total_rev else 0

        # Trend: compare first-half vs second-half monthly volumes
        mv = [pd["by_month"].get(k, 0) for k in months_sorted]
        if len(mv) >= 4:
            mid  = len(mv) // 2
            fh   = sum(mv[:mid]) / mid
            sh   = sum(mv[mid:]) / (len(mv) - mid)
            if   sh > fh * 1.1:  trend = "↑ Growing"
            elif sh < fh * 0.9:  trend = "↓ Declining"
            else:                 trend = "= Stable"
        elif len(mv) >= 2:
            trend = "↑ Growing" if mv[-1] >= mv[0] else "↓ Declining"
        else:
            trend = "— Insufficient data"

        ws13.append([pname, pd["cat"], pd["qty"], pd["revenue"],
                     avg_px, n_ord, round(avg_mo, 1), rs,
                     str(pd["last"]) if pd["last"] else "", trend])
        zar(ws13, ri, [4, 5])
        ws13.cell(ri, 8).number_format = PCT_FMT
        if "Growing"    in trend: paint(ws13, ri, 10, GREEN_F,  GREEN_T,  bold=True)
        elif "Declining" in trend: paint(ws13, ri, 10, RED_F,   RED_T,    bold=True)
        elif "Stable"    in trend: paint(ws13, ri, 10, AMBER_F, AMBER_T,  bold=True)

    zebra(ws13, 5, 4 + len(prod_sorted), len(h13))

    # ── Monthly units pivot ───────────────────────────────────────────────────
    ws13.append([]); ws13.append([])
    section(ws13, "UNITS SOLD BY PRODUCT × MONTH  (use for stock & reorder planning)")
    pivot_header_row = ws13.max_row

    ph = ["Product"] + [mlabel(monthly[k]["yr"], monthly[k]["mo"]) for k in months_sorted] + ["Total"]
    ws13.append(ph)
    for ci in range(1, len(ph) + 1):
        cell = ws13.cell(ws13.max_row, ci)
        cell.font  = Font(bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=HEADER2)
        cell.alignment = Alignment(horizontal="center")

    for pname, pd in prod_sorted:
        row = [pname] + [pd["by_month"].get(k, 0) for k in months_sorted] + [pd["qty"]]
        ws13.append(row)

    auto_width(ws13)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 14 — Customer Insights (RFM)
    # ══════════════════════════════════════════════════════════════════════════
    ws14 = wb.create_sheet("Customer Insights")
    ws14.append(["CUSTOMER INSIGHTS — RECENCY · FREQUENCY · VALUE"])
    ws14["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws14.append(["Ranked by lifetime revenue. "
                 "Status flags identify accounts that need attention before they lapse."])
    ws14.cell(2, 1).font = Font(italic=True, color="666666")
    ws14.append([])

    h14 = ["Customer", "# Orders", "Lifetime Revenue", "Avg Order Value",
           "First Order", "Last Order", "Days Since Last Order",
           "Orders/Month", "Revenue Rank", "Status"]
    ws14.append(h14); hdr(ws14, 4, len(h14))

    cust_sorted = sorted(cust.items(), key=lambda x: x[1]["revenue"], reverse=True)

    for ri, (cname, cd) in enumerate(cust_sorted, 5):
        days_since = (today - cd["last"]).days if cd["last"] else 9999
        span_mo = max(((cd["last"] - cd["first"]).days / 30.0), 1) \
                  if cd["first"] and cd["last"] else 1
        opm = cd["orders"] / span_mo
        avg = cd["revenue"] / cd["orders"] if cd["orders"] else 0

        if   cd["orders"] == 1 and days_since <= 60:
            status, fill, fc = "🆕 New",          PURPLE_F, PURPLE_T
        elif days_since <= 30:
            status, fill, fc = "✅ Active",         GREEN_F,  GREEN_T
        elif days_since <= 60:
            status, fill, fc = "🟡 Quiet",          AMBER_F,  AMBER_T
        elif days_since <= 90:
            status, fill, fc = "⚠ At Risk",         AMBER_F,  AMBER_T
        elif days_since <= 180:
            status, fill, fc = "🔴 Lapsing",         RED_F,    RED_T
        else:
            status, fill, fc = "💀 Lapsed",          RED_F,    RED_T

        ws14.append([cname, cd["orders"], cd["revenue"], avg,
                     str(cd["first"]) if cd["first"] else "",
                     str(cd["last"])  if cd["last"]  else "",
                     days_since, round(opm, 2),
                     f"#{ri - 4}", status])
        zar(ws14, ri, [3, 4])
        paint(ws14, ri, 10, fill, fc, bold=True)

    zebra(ws14, 5, 4 + len(cust_sorted), len(h14))

    # ── At-risk summary ───────────────────────────────────────────────────────
    ws14.append([]); ws14.append([])
    section(ws14, "AT-RISK SUMMARY")
    ws14.append(["Status", "Count", "Total Revenue at Risk"])
    hdr(ws14, ws14.max_row, 3, color=HEADER2)

    buckets = {
        "⚠ At Risk (60-90d)":   [c for c in cust_sorted if 60 < (today - c[1]["last"]).days <= 90  if c[1]["last"]],
        "🔴 Lapsing (90-180d)":  [c for c in cust_sorted if 90 < (today - c[1]["last"]).days <= 180 if c[1]["last"]],
        "💀 Lapsed (180d+)":     [c for c in cust_sorted if       (today - c[1]["last"]).days > 180  if c[1]["last"]],
    }
    for label, group in buckets.items():
        rev_risk = sum(c[1]["revenue"] for c in group)
        ws14.append([label, len(group), rev_risk])
        zar(ws14, ws14.max_row, [3])

    auto_width(ws14)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 14b — Best Customers (Top 20 by revenue)
    # ══════════════════════════════════════════════════════════════════════════
    wsbc = wb.create_sheet("Best Customers")
    wsbc.append(["BEST CUSTOMERS — TOP 20 BY LIFETIME REVENUE"])
    wsbc["A1"].font = Font(bold=True, size=16, color=PRIMARY)
    wsbc.append([f"Generated: {today.strftime('%d %B %Y')} · "
                 f"Based on {len(active)} active orders across {len(cust)} customers"])
    wsbc.cell(2, 1).font = Font(italic=True, color="666666")
    wsbc.append([])

    # Summary podium — top 3
    section(wsbc, "🏆 TOP 3 CUSTOMERS")
    podium = sorted(cust.items(), key=lambda x: x[1]["revenue"], reverse=True)[:3]
    medals = ["🥇", "🥈", "🥉"]
    wsbc.append(["Rank", "Customer", "Lifetime Revenue", "Orders", "Avg Order",
                 "Revenue Share", "Last Order", "Status"])
    hdr(wsbc, wsbc.max_row, 8, color=PRIMARY)
    total_cust_rev = sum(c["revenue"] for c in cust.values()) or 1
    for mi, (cname, cd) in enumerate(podium):
        days_since = (today - cd["last"]).days if cd["last"] else 9999
        avg = cd["revenue"] / cd["orders"] if cd["orders"] else 0
        share = cd["revenue"] / total_cust_rev
        if   cd["orders"] == 1 and days_since <= 60:  status = "🆕 New"
        elif days_since <= 30:                          status = "✅ Active"
        elif days_since <= 60:                          status = "🟡 Quiet"
        elif days_since <= 90:                          status = "⚠ At Risk"
        elif days_since <= 180:                         status = "🔴 Lapsing"
        else:                                           status = "💀 Lapsed"
        wsbc.append([f"{medals[mi]} #{mi+1}", cname, cd["revenue"], cd["orders"], avg,
                     share, str(cd["last"]) if cd["last"] else "", status])
        ri = wsbc.max_row
        zar(wsbc, ri, [3, 5])
        wsbc.cell(ri, 6).number_format = PCT_FMT
        # Gold / silver / bronze row highlight
        fill_colors = ["FFF4CC", "F2F2F2", "FCEBD9"]
        for c_idx in range(1, 9):
            wsbc.cell(ri, c_idx).fill = PatternFill("solid", fgColor=fill_colors[mi])
            wsbc.cell(ri, c_idx).font = Font(bold=True)

    # Full top-20 table
    wsbc.append([]); wsbc.append([])
    section(wsbc, "TOP 20 RANKED TABLE")
    h_bc = ["Rank", "Customer", "Lifetime Revenue", "Revenue Share",
            "# Orders", "Avg Order Value", "Orders/Month",
            "First Order", "Last Order", "Days Since Last", "Status"]
    wsbc.append(h_bc); hdr(wsbc, wsbc.max_row, len(h_bc))

    top20 = sorted(cust.items(), key=lambda x: x[1]["revenue"], reverse=True)[:20]
    for ri_idx, (cname, cd) in enumerate(top20, wsbc.max_row + 1):
        days_since = (today - cd["last"]).days if cd["last"] else 9999
        span_mo = max(((cd["last"] - cd["first"]).days / 30.0), 1) \
                  if cd["first"] and cd["last"] else 1
        opm = cd["orders"] / span_mo
        avg = cd["revenue"] / cd["orders"] if cd["orders"] else 0
        share = cd["revenue"] / total_cust_rev

        if   cd["orders"] == 1 and days_since <= 60:  status, fill, fc = "🆕 New",       PURPLE_F, PURPLE_T
        elif days_since <= 30:                          status, fill, fc = "✅ Active",     GREEN_F,  GREEN_T
        elif days_since <= 60:                          status, fill, fc = "🟡 Quiet",      AMBER_F,  AMBER_T
        elif days_since <= 90:                          status, fill, fc = "⚠ At Risk",    AMBER_F,  AMBER_T
        elif days_since <= 180:                         status, fill, fc = "🔴 Lapsing",    RED_F,    RED_T
        else:                                           status, fill, fc = "💀 Lapsed",     RED_F,    RED_T

        rank_num = ri_idx - (wsbc.max_row)  # row offset to rank
        wsbc.append([f"#{ri_idx - wsbc.max_row + 1}", cname, cd["revenue"], share,
                     cd["orders"], avg, round(opm, 2),
                     str(cd["first"]) if cd["first"] else "",
                     str(cd["last"])  if cd["last"]  else "",
                     days_since, status])
        row = wsbc.max_row
        zar(wsbc, row, [3, 6])
        wsbc.cell(row, 4).number_format = PCT_FMT
        paint(wsbc, row, 11, fill, fc, bold=True)

    # re-stripe (skip podium rows which are already coloured)
    data_start = wsbc.max_row - len(top20) + 1
    for r in range(data_start, wsbc.max_row + 1):
        if r % 2 == 0:
            for c_idx in range(1, len(h_bc)):  # skip status col which is already painted
                curr = wsbc.cell(r, c_idx).fill.fgColor.rgb
                if curr in ("00000000", "FFFFFFFF", "00FFFFFF"):
                    wsbc.cell(r, c_idx).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    # Revenue concentration chart data
    wsbc.append([]); wsbc.append([])
    section(wsbc, "REVENUE CONCENTRATION")
    wsbc.append(["Segment", "# Customers", "Revenue", "Revenue Share"])
    hdr(wsbc, wsbc.max_row, 4, color=HEADER2)
    all_sorted = sorted(cust.items(), key=lambda x: x[1]["revenue"], reverse=True)
    n_total = len(all_sorted)
    buckets_conc = [
        ("Top 1 customer",   all_sorted[:1]),
        ("Top 3 customers",  all_sorted[:3]),
        ("Top 5 customers",  all_sorted[:5]),
        ("Top 10 customers", all_sorted[:10]),
        ("All customers",    all_sorted),
    ]
    for label, group in buckets_conc:
        g_rev = sum(c[1]["revenue"] for c in group)
        share = g_rev / total_cust_rev
        wsbc.append([label, len(group), g_rev, share])
        ri = wsbc.max_row
        zar(wsbc, ri, [3])
        wsbc.cell(ri, 4).number_format = PCT_FMT
        if   share >= 0.5: paint(wsbc, ri, 4, AMBER_F, AMBER_T)
        elif share >= 0.8: paint(wsbc, ri, 4, RED_F,   RED_T)

    auto_width(wsbc)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 15 — Demand Forecast
    # ══════════════════════════════════════════════════════════════════════════
    ws15 = wb.create_sheet("Demand Forecast")
    ws15.append(["DEMAND FORECAST — 3-MONTH PROJECTION"])
    ws15["A1"].font = Font(bold=True, size=14, color=PRIMARY)

    confidence = "HIGH" if len(months_sorted) >= 9 else \
                 "MEDIUM" if len(months_sorted) >= 4 else "LOW"
    conf_color = GREEN_T if confidence == "HIGH" else \
                 AMBER_T if confidence == "MEDIUM" else RED_T

    ws15.append([f"Data: {len(months_sorted)} months · Confidence: {confidence} · "
                 f"Method: Weighted blend of linear regression (60%) + 3-month rolling average (40%)"])
    ws15.cell(2, 1).font = Font(italic=True, color="666666", size=10)
    ws15.append(["Projections are statistical estimates only. "
                 "Review alongside sales pipeline and known upcoming orders."])
    ws15.cell(3, 1).font = Font(italic=True, color="999999", size=9)
    ws15.append([])

    rev_fc  = forecast(rev_series)
    ord_fc  = forecast([monthly[k]["orders"] for k in months_sorted])
    item_fc = forecast([monthly[k]["items"]   for k in months_sorted])

    # ── Recent actuals ────────────────────────────────────────────────────────
    section(ws15, "RECENT ACTUALS (last 12 months)")
    h15a = ["Month", "# Orders", "Revenue", "Avg Order Value", "Units Sold"]
    ws15.append(h15a); hdr(ws15, ws15.max_row, len(h15a))
    for k in months_sorted[-12:]:
        m = monthly[k]
        avg = m["revenue"] / m["orders"] if m["orders"] else 0
        ws15.append([mlabel(m["yr"], m["mo"]), m["orders"], m["revenue"], avg, m["items"]])
        zar(ws15, ws15.max_row, [3, 4])
    ws15.append([])

    # ── 3-month projections ───────────────────────────────────────────────────
    section(ws15, "3-MONTH FORWARD PROJECTION")
    h15b = ["Period", "Projected Orders", "Projected Revenue",
            "Projected Avg Order", "Projected Units", "vs Last Month", "Confidence"]
    ws15.append(h15b); hdr(ws15, ws15.max_row, len(h15b), color=HEADER2)

    last_rev = rev_series[-1] if rev_series else 0
    if months_sorted:
        ly, lm = monthly[months_sorted[-1]]["yr"], monthly[months_sorted[-1]]["mo"]
    else:
        ly, lm = today.year, today.month

    for i in range(3):
        fy, fm  = add_months(ly, lm, i + 1)
        pr      = rev_fc[i]
        po      = max(round(ord_fc[i]), 0)
        pi      = max(round(item_fc[i]), 0)
        pa      = pr / po if po else 0
        vs_last = (pr - last_rev) / last_rev if last_rev else 0
        ws15.append([mlabel(fy, fm), po, pr, pa, pi, vs_last, confidence])
        ri = ws15.max_row
        zar(ws15, ri, [3, 4])
        ws15.cell(ri, 6).number_format = PCT_FMT
        cf, ct = ((GREEN_F, GREEN_T) if confidence == "HIGH" else
                  (AMBER_F, AMBER_T) if confidence == "MEDIUM" else
                  (RED_F,   RED_T))
        paint(ws15, ri, 7, cf, ct, bold=True)
        if vs_last > 0.05:  paint(ws15, ri, 6, GREEN_F, GREEN_T)
        elif vs_last < -0.05: paint(ws15, ri, 6, RED_F, RED_T)
        last_rev = pr

    ws15.append([])

    # ── Product-level demand forecast ─────────────────────────────────────────
    section(ws15, "PRODUCT-LEVEL DEMAND FORECAST (units/month)")
    h15c = ["Product", "Avg Units/Month (actual)",
            f"Forecast {mlabel(*add_months(ly, lm, 1))}",
            f"Forecast {mlabel(*add_months(ly, lm, 2))}",
            f"Forecast {mlabel(*add_months(ly, lm, 3))}",
            "Trend"]
    ws15.append(h15c); hdr(ws15, ws15.max_row, len(h15c), color=HEADER2)

    for pname, pd in prod_sorted[:20]:          # top 20 products
        mv = [pd["by_month"].get(k, 0) for k in months_sorted]
        pfc = forecast(mv)
        avg_act = sum(mv) / n_months
        mv_last = mv[-1] if mv else 0
        projected_last = pfc[2]
        trend = ("↑" if projected_last > avg_act * 1.05 else
                 "↓" if projected_last < avg_act * 0.95 else "=")
        ws15.append([pname, round(avg_act, 1),
                     round(pfc[0], 1), round(pfc[1], 1), round(pfc[2], 1), trend])
        ri = ws15.max_row
        if trend == "↑": paint(ws15, ri, 6, GREEN_F, GREEN_T, bold=True)
        elif trend == "↓": paint(ws15, ri, 6, RED_F, RED_T, bold=True)

    ws15.append([])

    # ── Seasonal index ────────────────────────────────────────────────────────
    if len(months_sorted) >= 6:
        section(ws15, "SEASONAL CALENDAR INDEX  (index > 1.0 = above-average demand month)")
        ws15.append(["Use to plan stock levels, staffing and marketing spend by time of year."])
        ws15.cell(ws15.max_row, 1).font = Font(italic=True, color="666666", size=9)
        ws15.append([])
        h15d = ["Calendar Month", "Avg Revenue", "Seasonal Index",
                "Seasons Observed", "Signal"]
        ws15.append(h15d); hdr(ws15, ws15.max_row, len(h15d), color=HEADER2)
        overall_avg = sum(rev_series) / len(rev_series)
        cal = defaultdict(list)
        for k in months_sorted:
            cal[monthly[k]["mo"]].append(monthly[k]["revenue"])
        for mo in range(1, 13):
            vals = cal.get(mo, [])
            if vals:
                avg_r = sum(vals) / len(vals)
                idx   = avg_r / overall_avg if overall_avg else 1
                if   idx >= 1.15: sig, sf, st = "🔥 Peak — stock up",    GREEN_F, GREEN_T
                elif idx >= 0.95: sig, sf, st = "✓ Normal",               "FFFFFF", "000000"
                elif idx >= 0.80: sig, sf, st = "↓ Below average",        AMBER_F, AMBER_T
                else:             sig, sf, st = "❄ Slow — reduce costs",  RED_F,   RED_T
                ws15.append([MONTH_NAMES[mo - 1], avg_r, round(idx, 2), len(vals), sig])
                ri = ws15.max_row
                ws15.cell(ri, 2).number_format = ZAR_FMT
                paint(ws15, ri, 5, sf, st)
            else:
                ws15.append([MONTH_NAMES[mo - 1], "No data", "—", 0, "No data yet"])

    auto_width(ws15)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 16 — Delivery Performance
    # ══════════════════════════════════════════════════════════════════════════
    ws16 = wb.create_sheet("Delivery Performance")
    ws16.append(["DELIVERY PERFORMANCE ANALYSIS"])
    ws16["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws16.append([])

    completed_d = [d for d in deliveries if d.status == "completed" and d.actual_date]
    on_time_d   = [d for d in completed_d if d.actual_date <= d.scheduled_date]
    delays      = [(d.actual_date - d.scheduled_date).days
                   for d in completed_d if d.actual_date > d.scheduled_date]
    otr         = len(on_time_d) / len(completed_d) if completed_d else 0
    avg_delay   = sum(delays) / len(delays) if delays else 0

    section(ws16, "KEY PERFORMANCE INDICATORS")
    ws16.append(["Metric", "Value"])
    hdr(ws16, ws16.max_row, 2)
    kpis = [
        ("Total Deliveries",              len(deliveries)),
        ("Completed",                     len(completed_d)),
        ("Pending / Scheduled",           len([d for d in deliveries if d.status != "completed"])),
        ("On-Time Rate",                  f"{otr:.1%}"),
        ("Late Deliveries",               len(delays)),
        ("Avg Delay (days, when late)",   round(avg_delay, 1)),
        ("Customer Deliveries",           len([d for d in deliveries if (d.delivery_type or "customer") == "customer"])),
        ("Hardware Deliveries",           len([d for d in deliveries if d.delivery_type == "hardware"])),
    ]
    for ki, (k, v) in enumerate(kpis):
        ws16.append([k, v])
        if "Rate" in k and isinstance(v, str) and "%" in v and float(v.replace("%","")) >= 90:
            paint(ws16, ws16.max_row, 2, GREEN_F, GREEN_T)
        elif "Rate" in k and isinstance(v, str) and "%" in v:
            paint(ws16, ws16.max_row, 2, RED_F, RED_T)
    zebra(ws16, ws16.max_row - len(kpis) + 1, ws16.max_row, 2)

    # ── Hardware Store Breakdown ──────────────────────────────────────────────
    hw_deliveries = [d for d in deliveries if d.delivery_type == "hardware"]
    if hw_deliveries:
        ws16.append([]); ws16.append([])
        section(ws16, "HARDWARE STORE BREAKDOWN")
        ws16.append(["Hardware Store", "Total Deliveries", "Completed", "Pending", "On-Time Rate"])
        hdr(ws16, ws16.max_row, 5, color=HEADER2)

        hw_stores = defaultdict(lambda: {"total": 0, "done": 0, "ot": 0})
        for d in hw_deliveries:
            store = getattr(d, "hardware_store_name", None) or "Unspecified"
            hw_stores[store]["total"] += 1
            if d.status == "completed":
                hw_stores[store]["done"] += 1
                if d.actual_date and d.actual_date <= d.scheduled_date:
                    hw_stores[store]["ot"] += 1

        for store, sd in sorted(hw_stores.items(), key=lambda x: x[1]["total"], reverse=True):
            pending = sd["total"] - sd["done"]
            ot_rate = sd["ot"] / sd["done"] if sd["done"] else 0
            ws16.append([store, sd["total"], sd["done"], pending,
                         f"{ot_rate:.1%}" if sd["done"] else "N/A"])
            ri = ws16.max_row
            if sd["done"] > 0:
                paint(ws16, ri, 5,
                      GREEN_F if ot_rate >= 0.9 else AMBER_F if ot_rate >= 0.7 else RED_F,
                      GREEN_T if ot_rate >= 0.9 else AMBER_T if ot_rate >= 0.7 else RED_T)
        zebra(ws16, ws16.max_row - len(hw_stores) + 1, ws16.max_row, 5)

    # Monthly delivery volume by type
    ws16.append([]); ws16.append([])
    section(ws16, "MONTHLY DELIVERIES BY TYPE")
    ws16.append(["Month", "Customer", "Hardware", "Total", "Completed", "On-Time Rate"])
    hdr(ws16, ws16.max_row, 6, color=HEADER2)
    del_monthly = defaultdict(lambda: {"cust": 0, "hw": 0, "done": 0, "ot": 0})
    for d in deliveries:
        mk = d.scheduled_date.strftime("%Y-%m")
        if (d.delivery_type or "customer") == "customer": del_monthly[mk]["cust"] += 1
        else:                                              del_monthly[mk]["hw"]   += 1
        if d.status == "completed":
            del_monthly[mk]["done"] += 1
            if d.actual_date and d.actual_date <= d.scheduled_date:
                del_monthly[mk]["ot"] += 1
    for mk in sorted(del_monthly.keys()):
        dm = del_monthly[mk]
        total_dm = dm["cust"] + dm["hw"]
        ot_rate  = dm["ot"] / dm["done"] if dm["done"] else 0
        yr, mo = int(mk[:4]), int(mk[5:])
        ws16.append([mlabel(yr, mo), dm["cust"], dm["hw"], total_dm, dm["done"], ot_rate])
        ri = ws16.max_row
        ws16.cell(ri, 6).number_format = PCT_FMT
        paint(ws16, ri, 6, GREEN_F if ot_rate >= 0.9 else AMBER_F if ot_rate >= 0.7 else RED_F,
              GREEN_T if ot_rate >= 0.9 else AMBER_T if ot_rate >= 0.7 else RED_T)

    # Detail table
    ws16.append([]); ws16.append([])
    section(ws16, "DELIVERY DETAIL")
    h16d = ["Order #", "Customer", "Type", "Hardware Store", "Scheduled", "Actual",
            "Day Delta", "Status", "On Time?"]
    ws16.append(h16d); hdr(ws16, ws16.max_row, len(h16d))
    for d in sorted(deliveries, key=lambda x: x.scheduled_date, reverse=True):
        delta = ""
        flag  = "⏳ Pending"
        ff, ft = GRAY_F, "666666"
        if d.actual_date:
            delta = (d.actual_date - d.scheduled_date).days
            if delta <= 0:
                flag, ff, ft = "✓ On time", GREEN_F, GREEN_T
            else:
                flag, ff, ft = f"✗ {delta}d late", RED_F, RED_T
        ws16.append([d.order.order_number, d.order.customer.name,
                     d.delivery_type or "customer",
                     getattr(d, "hardware_store_name", None) or "",
                     str(d.scheduled_date),
                     str(d.actual_date) if d.actual_date else "Pending",
                     delta, d.status, flag])
        paint(ws16, ws16.max_row, 9, ff, ft)

    auto_width(ws16)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 17 — TLB Utilisation
    # ══════════════════════════════════════════════════════════════════════════
    ws17 = wb.create_sheet("TLB Utilisation")
    ws17.append(["TLB UTILISATION & REVENUE TRENDS"])
    ws17["A1"].font = Font(bold=True, size=14, color=PRIMARY)
    ws17.append([])

    total_hrs    = sum(t.hours_billed for t in tlbs)
    total_tlb_r  = sum(t.total_zar    for t in tlbs)

    section(ws17, "SUMMARY")
    ws17.append(["Metric", "Value"])
    hdr(ws17, ws17.max_row, 2)
    tlb_kpis = [
        ("Total TLB Bookings",      len(tlbs)),
        ("Total Hours Billed",      round(total_hrs, 1)),
        ("Total TLB Revenue",       total_tlb_r),
        ("Avg Hours per Booking",   round(total_hrs / len(tlbs), 1) if tlbs else 0),
        ("Avg Revenue per Booking", total_tlb_r / len(tlbs) if tlbs else 0),
        ("Avg Revenue per Hour",    total_tlb_r / total_hrs if total_hrs else 0),
    ]
    for ki, (k, v) in enumerate(tlb_kpis):
        ws17.append([k, v])
        if isinstance(v, float) and "Revenue" in k:
            ws17.cell(ws17.max_row, 2).number_format = ZAR_FMT
    zebra(ws17, ws17.max_row - len(tlb_kpis) + 1, ws17.max_row, 2)

    # Monthly trend
    ws17.append([]); ws17.append([])
    section(ws17, "MONTHLY TLB TREND")
    h17 = ["Month", "Bookings", "Hours Billed", "Revenue",
           "Avg Hours/Booking", "Revenue/Hour", "MoM Revenue Δ%"]
    ws17.append(h17); hdr(ws17, ws17.max_row, len(h17), color=HEADER2)
    tlb_m = defaultdict(lambda: {"n": 0, "hrs": 0.0, "rev": 0.0})
    for t in tlbs:
        mk = t.booking_date.strftime("%Y-%m")
        tlb_m[mk]["n"]   += 1
        tlb_m[mk]["hrs"] += t.hours_billed
        tlb_m[mk]["rev"] += t.total_zar
    prev_tlb_rev = None
    for mk in sorted(tlb_m.keys()):
        tm = tlb_m[mk]
        yr, mo = int(mk[:4]), int(mk[5:])
        avg_h = tm["hrs"] / tm["n"] if tm["n"] else 0
        rph   = tm["rev"] / tm["hrs"] if tm["hrs"] else 0
        mom   = (tm["rev"] - prev_tlb_rev) / prev_tlb_rev if prev_tlb_rev else ""
        ws17.append([mlabel(yr, mo), tm["n"], round(tm["hrs"], 1), tm["rev"],
                     round(avg_h, 1), round(rph, 2),
                     mom])
        ri = ws17.max_row
        zar(ws17, ri, [4, 6])
        if mom != "":
            ws17.cell(ri, 7).number_format = PCT_FMT
            if   mom >  0.1: paint(ws17, ri, 7, GREEN_F, GREEN_T)
            elif mom < -0.1: paint(ws17, ri, 7, RED_F,   RED_T)
        prev_tlb_rev = tm["rev"]

    # Customer breakdown
    ws17.append([]); ws17.append([])
    section(ws17, "TOP TLB CUSTOMERS")
    h17b = ["Customer", "Bookings", "Total Hours", "Total Revenue",
            "Avg Hours/Booking", "Revenue Share"]
    ws17.append(h17b); hdr(ws17, ws17.max_row, len(h17b), color=HEADER2)
    tlb_c = defaultdict(lambda: {"n": 0, "hrs": 0.0, "rev": 0.0})
    for t in tlbs:
        tlb_c[t.customer.name]["n"]   += 1
        tlb_c[t.customer.name]["hrs"] += t.hours_billed
        tlb_c[t.customer.name]["rev"] += t.total_zar
    for cname, cd in sorted(tlb_c.items(), key=lambda x: x[1]["rev"], reverse=True):
        avg_h = cd["hrs"] / cd["n"] if cd["n"] else 0
        rs    = cd["rev"] / total_tlb_r if total_tlb_r else 0
        ws17.append([cname, cd["n"], round(cd["hrs"], 1), cd["rev"], round(avg_h, 1), rs])
        ri = ws17.max_row
        zar(ws17, ri, [4])
        ws17.cell(ri, 6).number_format = PCT_FMT

    # TLB 3-month forecast
    ws17.append([]); ws17.append([])
    section(ws17, "TLB DEMAND FORECAST (3 months)")
    h17c = ["Period", "Projected Bookings", "Projected Hours", "Projected Revenue"]
    ws17.append(h17c); hdr(ws17, ws17.max_row, len(h17c), color=HEADER2)
    tlb_keys = sorted(tlb_m.keys())
    tlb_rev_s = [tlb_m[k]["rev"] for k in tlb_keys]
    tlb_hrs_s = [tlb_m[k]["hrs"] for k in tlb_keys]
    tlb_n_s   = [tlb_m[k]["n"]   for k in tlb_keys]
    tlb_rev_f = forecast(tlb_rev_s)
    tlb_hrs_f = forecast(tlb_hrs_s)
    tlb_n_f   = forecast(tlb_n_s)
    tly, tlmo = ((monthly[months_sorted[-1]]["yr"], monthly[months_sorted[-1]]["mo"])
                  if months_sorted else (today.year, today.month))
    for i in range(3):
        fy, fm = add_months(tly, tlmo, i + 1)
        ws17.append([mlabel(fy, fm),
                     max(round(tlb_n_f[i]), 0),
                     round(max(tlb_hrs_f[i], 0), 1),
                     max(tlb_rev_f[i], 0)])
        ri = ws17.max_row
        zar(ws17, ri, [4])
        for c in range(1, 5):
            ws17.cell(ri, c).fill = PatternFill("solid", fgColor="F0F4FF")

    auto_width(ws17)

    # ── Save ──────────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
